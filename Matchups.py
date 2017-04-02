import requests
from openpyxl import Workbook

wb = Workbook()

standingsUrl = "http://data.nba.com/data/10s/v2015/json/mobile_teams/nba/2015/scores/00_playoff_bracket.json"

js = requests.get(standingsUrl, headers={"referer": "http://stats.nba.com/playoffs/",
                                         "User-Agent": "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.87 Safari/537.36"})
response = js.json()["pb"]["r"][0]["co"]

matchups = {}

for conference in range(2):
    for i in range(0, 4):
        team1ID = response[conference]["ser"][i]["tid1"]
        team2ID = response[conference]["ser"][i]["tid2"]
        team1Name = response[conference]["ser"][i]["tc1"]
        team2Name = response[conference]["ser"][i]["tc2"]
        matchups[(team1ID, team2ID)] = (team1Name, team2Name)

print(matchups)
num = 2
ws = wb.active
ws['A1'] = "Name"
ws['B1'] = "Team"
ws['C1'] = "Minutes Per Game"
ws['D1'] = "PIE Difference"
ws['E1'] = "Net Rating Difference"

for match in matchups.keys():
    for i in range(0, 2):
        currentTeam = match[i]
        otherTeam = match[1-i]

        players = {}
        playersVsMatchUrl = "http://stats.nba.com/stats/leaguedashplayerstats?College=&Conference=&Country=&DateFrom" \
                             "=&DateTo=&Division=&DraftPick=&DraftYear=&GameScope=&GameSegment=&Height=&LastNGames=0" \
                             "&LeagueID=00&Location=&MeasureType=Advanced&Month=0&OpponentTeamID=" + str(otherTeam) + "&" \
                             "Outcome=&PORound=0&PaceAdjust=N&PerMode=Totals&Period=0&PlayerExperience=&Player" \
                             "Position=&PlusMinus=N&Rank=N&Season=2015-16&SeasonSegment=&SeasonType=Regular+Season&Sh" \
                             "otClockRange=&StarterBench=&TeamID=" + str(currentTeam) + "&VsConference=&VsDivision=&Weight="
        js = requests.get(playersVsMatchUrl, headers={"referer": "http://stats.nba.com/player/",
                                                      "User-Agent": "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.87 Safari/537.36"})
        playersVsMatch = js.json()["resultSets"][0]["rowSet"]
        for p in playersVsMatch:
            players[p[0]] = [p[1], p[12], p[24], p[9]]

        playersVsLeagueUrl = "http://stats.nba.com/stats/leaguedashplayerstats?College=&Conference=&Country=&DateFrom" \
                             "=&DateTo=&Division=&DraftPick=&DraftYear=&GameScope=&GameSegment=&Height=&LastNGames=0" \
                             "&LeagueID=00&Location=&MeasureType=Advanced&Month=0&OpponentTeamID=0&" \
                             "Outcome=&PORound=0&PaceAdjust=N&PerMode=Totals&Period=0&PlayerExperience=&Player" \
                             "Position=&PlusMinus=N&Rank=N&Season=2015-16&SeasonSegment=&SeasonType=Regular+Season&Sh" \
                             "otClockRange=&StarterBench=&TeamID=" + str(currentTeam) + "&VsConference=&VsDivision=&Weight="
        js = requests.get(playersVsLeagueUrl, headers={"referer": "http://stats.nba.com/player/",
                                                       "User-Agent": "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.87 Safari/537.36"})
        playersVsLeague = js.json()["resultSets"][0]["rowSet"]
        for p in playersVsLeague:
            if p[0] in players.keys():
                players[p[0]][1] -= p[12]
                players[p[0]][2] -= p[24]

        for p in players.values():
            ws['A' + str(num)] = p[0]
            ws['B' + str(num)] = matchups[match][i]
            ws['C' + str(num)] = p[3]
            ws['D' + str(num)] = round(p[2], 2)
            ws['E' + str(num)] = round(p[1], 2)
            num += 1

wb.save('C:\\Users\\tuvia\\Documents\\Analytics\\Playoff Matchups.xlsx')
