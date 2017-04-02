
# Returns the score differential in the given possession.
# _index -- index of current possession.
def find_difference(_index):
    while playByPlay[_index][10] is None:
        if playByPlay[_index][2] == 12:
            return 0
        _index -= 1

    if playByPlay[_index][11] == "TIE":
        diff = 0
    else:
        diff = int(playByPlay[_index][11]) if play[12] == 2 else int(playByPlay[_index][11]) * -1
    return diff


# Counts given number of possessions from given index and returns new index.
# _index -- starting index.
# num -- number of possessions to count.
def find_x_possessions(_index, num):
    free_throws = [[12, 15] if num > 0 else [11, 13]][0]
    possessions = 0
    _index += int(copysign(1, num))
    next_is_offensive_rebound = False
    while (possessions != abs(num) * 2 or next_is_offensive_rebound) and _index != len(playByPlay) - 1 and playByPlay[_index][2] not in [9, 12]:
        is_free_throw = playByPlay[_index][2] == 3
        if playByPlay[_index][2] in [1, 2, 5] or (is_free_throw and playByPlay[_index][3] in free_throws):
            possessions += 1
        if next_is_offensive_rebound:
            next_is_offensive_rebound = False
            possessions -= 1
        next_is_offensive_rebound = playByPlay[_index + 1][2] == 4 and not (is_free_throw and playByPlay[_index][3] in [11, 13]) and \
                                 playByPlay[_index][15] in playByPlay[_index + 1]
        _index += int(copysign(1, num))

    if num < 0:
        not_used = 0
        original_index = _index + 1
        while playByPlay[_index][2] not in range(1, 6) or playByPlay[original_index][15] == playByPlay[_index][15]:
            if playByPlay[_index][2] not in [1, 2, 3, 5]:
                not_used += 1
            else:
                not_used = 0
            _index -= 1
        return _index + not_used

    extra = 0
    while playByPlay[_index + extra][2] in [6, 8]:
        extra += 1
    if playByPlay[_index + extra][2] == 3 and playByPlay[_index + extra][3] == 10:
        return _index + extra
    return _index - 1


# Returns the points scored by the given team until the given possession.
# _index -- index of current possession.
# team -- team number (4 for home team or 5 for away team)
def find_points(_index, team):
    while playByPlay[_index][10] is None:
        if playByPlay[_index][2] == 12:
            return 0
        _index -= 1
    a = playByPlay[_index][10].index("-")
    if team == 5:
        return int(playByPlay[_index][10][:a - 1])
    return int(playByPlay[_index][10][a + 2:])


# Returns the amount of points the opponent of the given team has scored in a row.
# _index -- index of current possession.
# team -- team number (4 for home team or 5 for away team)
def find_opponent_streak(_index, team):
    original_index = _index
    while not (playByPlay[_index][10] is not None and playByPlay[_index][12] == team) and _index != 0:
        _index -= 1
    team = int(not (team - 4)) + 4
    if _index == 0:
        return find_points(original_index, team)
    return find_points(original_index, team) - find_points(_index, team)


# Returns the number of possessions from the last or to the next timeout (or end of period).
# _index -- index of current possession.
# direction -- 1 in order to count to the next timeout, -1 in order to count to the previous timeout.
def find_possessions_from_last_timeout(_index, direction):
    possessions = 0
    while playByPlay[_index][2] not in [9, 12, 13]:
        if playByPlay[_index][2] in [1, 2, 5, 7] or playByPlay[_index][2] == 3 and playByPlay[_index][3] in [10, 12, 15]:
            possessions += 1
        if playByPlay[_index][2] == 4 and playByPlay[_index][15] == playByPlay[_index - 1][15]:
            possessions -= 1

        _index += direction
    return possessions / 2


def handle_remaining_timeouts(timeouts):
    for x in timeouts.values():
        if play[4] > 3 and x[0] > 3:
            x[0] = 3
        if play[4] > 4:
            x[0] = 3
    return timeouts


import requests
from openpyxl import Workbook
from math import copysign

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.99 Safari/537.36"}

a = ["Team", "Game ID", "Date", "Wins", "Losses", "Period", "Minutes", "Remaining Timeouts", "Score Difference",
     "Lead Or Deficit", "Opponent Streak", "Possessions From Previous Timeout", "Possessions To Next Timeout",
     "Difference In Last 5 Possessions", "Points In Last 5 Possessions", "-5", "-4", "-3", "-2", "-1",
     "Points In Next 5 Possessions", "1", "2", "3", "4", "5", "Difference In Next 5 Possessions",
     "Last possession description"]

wb = Workbook()
ws = wb.active
try:
    for x in range(len(a)):
        ws[chr(x + 65) + '1'] = a[x]
except Exception:
    pass
ws['AA1'] = "Difference In Next 5 Possessions"
ws['AB1'] = "Last possession description"

row = 2

teamsUrl = "http://stats.nba.com/stats/leaguestandingsv3?LeagueID=00&Season=2015-16&SeasonType=Regular+Season"
teams = {x[2]: x[3] + " " + x[4] for x in requests.get(teamsUrl, headers=headers).json()["resultSets"][0]["rowSet"]}

for season in range(5, 6):
    for game in ["0021" + str(season) + "0" + str(x).zfill(4) for x in range(1, 1231)]:
        print(game)
        infoUrl = "http://stats.nba.com/stats/boxscoresummaryv2?GameID=" + game
        info = requests.get(infoUrl, headers=headers).json()["resultSets"][5]["rowSet"]
        home, away, win = info[0][7], info[1][7], info[0][22] > info[1][22]
        timeouts = {info[0][3]: [7, int(home[:home.index("-")]) - win, int(home[home.index("-") + 1:]) - (not win)],
                    info[1][3]: [7, int(away[:away.index("-")]) - (not win), int(away[away.index("-") + 1:]) - win]}
        date = requests.get(infoUrl, headers=headers).json()["resultSets"][4]["rowSet"][0][0]

        url = "http://stats.nba.com/stats/playbyplayv2?EndPeriod=10&EndRange=55800&GameID=" + game + "&RangeType=2&Season=" \
              "2015-16&SeasonType=Regular+Season&StartPeriod=1&StartRange=0"

        playByPlay = requests.get(url, headers=headers).json()["resultSets"][0]["rowSet"]
        for play in playByPlay:
            if play[2] == 9 and play[13]:
                try:
                    ws['A' + str(row)] = teams[play[13]]
                    ws['B' + str(row)] = int(game[6:])  # 2010 + season
                    ws['C' + str(row)] = date
                    ws['D' + str(row)] = timeouts[play[13]][1]
                    ws['E' + str(row)] = timeouts[play[13]][2]
                    ws['F' + str(row)] = play[4]
                    ws['G' + str(row)] = round(int(play[6][:play[6].index(":")]) + int(play[6][play[6].index(":") + 1:]) / 60, 2)
                    if play[4] > 3 or play[3] == 1:
                        timeouts[play[13]][0] -= 1

                    ws['H' + str(row)] = timeouts[play[13]][0]
                    i, t = playByPlay.index(play), play[12] + 2
                    gameDifference = find_difference(i)
                    ws['I' + str(row)] = gameDifference
                    ws['J' + str(row)] = int(gameDifference > 0) * 2 + int(not gameDifference)
                    ws['K' + str(row)] = find_opponent_streak(i, t)
                    ws['L' + str(row)] = find_possessions_from_last_timeout(i - 1, -1)
                    ws['M' + str(row)] = find_possessions_from_last_timeout(i + 1, 1)
                    ws['N' + str(row)] = gameDifference - find_difference(find_x_possessions(i, -5))
                    ws['O' + str(row)] = find_points(i, t) - find_points(find_x_possessions(i, -5), t)
                    for x in range(-4, 0):
                        ws[chr(84 + x) + str(row)] = find_points(find_x_possessions(i, x), t) - find_points(find_x_possessions(i, x - 1), t)
                    ws['T' + str(row)] = find_points(i, t) - find_points(find_x_possessions(i, -1), t)
                    ws['U' + str(row)] = find_points(find_x_possessions(i, 5), t) - find_points(i, t)
                    ws['V' + str(row)] = find_points(find_x_possessions(i, 1), t) - find_points(i, t)
                    for x in range(2, 6):
                        ws[chr(x + 85) + str(row)] = find_points(find_x_possessions(i, x), t) - find_points(find_x_possessions(i, x - 1), t)
                    ws['AA' + str(row)] = find_difference(find_x_possessions(i, 5)) - gameDifference
                    ws['AB' + str(row)] = ' '.join(filter(None, playByPlay[find_x_possessions(i, -0.5) + 1][7:10]))
                    row += 1
                except KeyError:
                    print(play[7], play[8], play[9])
                    continue

            if play[2] == 12:
                timeouts = handle_remaining_timeouts(timeouts)

wb.save('C:\\Users\\tuvia\\Documents\\Analytics\\Timeout Variables.xlsx')
