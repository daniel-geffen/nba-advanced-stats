def get_length(a):
    a = a.getText()
    feet = int(a[0])
    inches = float(a[3:len(a) - 1])
    inches += feet * 12
    return inches


def int_to_string(num):
    if num < 10:
        return "0" + str(num)
    else:
        return str(num)


from bs4 import BeautifulSoup
from urllib.request import urlopen
import requests
from openpyxl import Workbook
from sklearn.linear_model import LinearRegression

wb = Workbook()
ws = wb.active
ws['A1'] = "Name"
ws['B1'] = "Height"
ws['C1'] = "Wingspan"
ws['D1'] = "Ratio"
ws['E1'] = "Rookie"
ws['F1'] = "Sophomore"
ws['G1'] = "Year 3"
ws['H1'] = "Year 4"
ws['I1'] = "Year 5"

url = "http://www.draftexpress.com/nba-pre-draft-measurements.php?page=&year=All&source=All&sort2=DESC&draft=100&" \
      "pos=0&sort=5"

html1 = urlopen(url)

parsed = BeautifulSoup(html1, "html.parser")

players = {}

for row in parsed.find_all('tr')[1:]:
    try:
        l = row.find_all('td')
        d = l[0].getText()
        name = d[:d.index("-") - 1]
        year = int(d[d.index("-") + 2:])
        height = get_length(l[1])
        wingspan = get_length(l[4])
        # print(name, year, height, wingspan)
        if year >= 2000:
            if name not in players.keys():
                players[name] = {"Year": year,
                                 "Height": height,
                                 "Wingspan": wingspan}
            else:
                if players[name]["Year"] < year:
                    players[name]["Height"] = height
                    players[name]["Wingspan"] = wingspan

    except (UnicodeEncodeError, ValueError, IndexError):
        pass


NBA = {}
experiences = ("Rookie", "Sophomore", "Veteran")
years = {}
headers = {"referer": "http://stats.nba.com/league/player/",
           "User-Agent": "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.87 Safari/537.36"}

for y in range(16):
    print(y)
    y1 = int_to_string(y)
    y2 = int_to_string(y + 1)
    yearsUrl = "http://stats.nba.com/stats/leaguedashplayerbiostats?College=&Conference=&Country=&DateFrom=&DateTo=" \
               "&Division=&DraftPick=&DraftYear=&GameScope=&GameSegment=&Height=&LastNGames=0&LeagueID=00&Location=" \
               "&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PerMode=PerGame&Period=0&PlayerExperience=&PlayerPosit" \
               "ion=&Season=20" + y1 + "-" + y2 + "&SeasonSegment=&SeasonType=Regular+Season&ShotClockRange=&Starte" \
               "rBench=&TeamID=0&VsConference=&VsDivision=&Weight="
    response = requests.get(yearsUrl, headers=headers)
    js = response.json()["resultSets"][0]["rowSet"]
    for p in js:
        if p[1] not in years.keys() and p[10] != "Undrafted" and int(p[10]) >= 2000:
            years[p[1]] = {"Draft": p[10], "Pick": p[12]}

    for e in experiences:
        nbaUrl = "http://stats.nba.com/stats/leaguedashplayerstats?College=&Conference=&Country=&DateFrom=&DateTo=&Di" \
                 "vision=&DraftPick=&DraftYear=&GameScope=&GameSegment=&Height=&LastNGames=0&LeagueID=00&Location=&Me" \
                 "asureType=Base&Month=0&OpponentTeamID=0&Outcome=&PORound=0&PaceAdjust=N&PerMode=PerGame&Period=0&Pl" \
                 "ayerExperience=" + e + "&PlayerPosition=&PlusMinus=N&Rank=N&Season=20" + y1 + "-" + y2 + "&SeasonSe" \
                 "gment=&SeasonType=Regular+Season&ShotClockRange=&StarterBench=&TeamID=0&VsConference=&VsDivision=&Weight="
        response = requests.get(nbaUrl, headers=headers)
        js = response.json()["resultSets"][0]["rowSet"]

        for p in js:
            if p[9] >= 8:
                if p[1] not in NBA.keys():
                    NBA[p[1]] = {"Rookie": None,
                                 "Sophomore": None,
                                 "Veteran": []}
                if NBA[p[1]][e] is None:
                    NBA[p[1]][e] = p[15]
                elif type(NBA[p[1]][e]) == list:
                    NBA[p[1]][e].append(p[15])

for k in list(NBA):
    if NBA[k]["Rookie"] is None or NBA[k]["Sophomore"] is None:
        del NBA[k]


drafts = {}
for y in range(16):
    drafts["20" + int_to_string(y)] = []
for k, y in years.items():
    try:
        drafts[y["Draft"]].append((int(y["Pick"]), players[k]["Wingspan"]))
    except (KeyError, TypeError):
        pass
for k, d in drafts.items():
    suma = 0
    leng = 0
    for t in d:
        suma += t[1] * (31 - t[0])
        leng += 31 - t[0]
    drafts[k] = suma / leng

ws2 = wb.create_sheet(title="Drafts")
ws2['A1'] = "Year"
ws2['B1'] = "Average"
wsCounter = 2

for k, v in drafts.items():
    ws2['A' + str(wsCounter)] = int(k)
    ws2['B' + str(wsCounter)] = round(v, 2)
    wsCounter += 1

ws3 = wb.create_sheet(title="Multiple Regression")
ws3Counter = 2
wsCounter = 2
for a in list(players):
    try:
        players[a].update(NBA[a])
        ws['A' + str(wsCounter)] = a
        ws['B' + str(wsCounter)] = round(players[a]["Height"], 2)
        ws['C' + str(wsCounter)] = round(players[a]["Wingspan"], 2)
        ws['D' + str(wsCounter)] = round(players[a]["Wingspan"] / players[a]["Height"], 2)
        ws['E' + str(wsCounter)] = round(players[a]["Rookie"], 2)
        ws['F' + str(wsCounter)] = round(players[a]["Sophomore"], 2)
        ws['G' + str(wsCounter)] = round(players[a]["Veteran"][0], 2)
        ws['H' + str(wsCounter)] = round(players[a]["Veteran"][1], 2)
        ws['I' + str(wsCounter)] = round(players[a]["Veteran"][2], 2)

        wsCounter += 1

    except KeyError:
        del players[a]
        pass
    except (TypeError, TypeError, IndexError):
        wsCounter += 1
        pass

wb.save('C:\\Users\\tuvia\\Documents\\Analytics\\Wingspan.xlsx')
