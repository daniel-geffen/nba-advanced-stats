import requests
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws["A1"] = "Team"
ws["B1"] = "Minute Difference"

rowCounter = 2

teams = {}

teamsUrl = "http://stats.nba.com/stats/leaguedashteamstats?Conference=&DateFrom=&DateTo=&Division=&GameScope=&GameSeg" \
           "ment=&LastNGames=0&LeagueID=00&Location=&MeasureType=Base&Month=0&OpponentTeamID=0&Outcome=&PORound=0&Pac" \
           "eAdjust=N&PerMode=PerGame&Period=0&PlayerExperience=&PlayerPosition=&PlusMinus=N&Rank=N&Season=2015-16&Se" \
           "asonSegment=&SeasonType=Playoffs&ShotClockRange=&StarterBench=&TeamID=0&VsConference=&VsDivision="

response = requests.get(teamsUrl, headers={"referer": "http://stats.nba.com/team/"}).json()["resultSets"][0]["rowSet"]

for team in response:
    teams[(team[0], team[1])] = {}

print(teams)

heads = {"Referer": "http://stats.nba.com/team/",
         "User-Agent": "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.112 Safari/537.36"}

for team in teams.keys():
    playoffUrl = "http://stats.nba.com/stats/teamplayerdashboard?DateFrom=&DateTo=&GameSegment=&LastNGames=0&LeagueID" \
                 "=00&Location=&MeasureType=Base&Month=0&OpponentTeamID=0&Outcome=&PaceAdjust=N&PerMode=PerGame&Perio" \
                 "d=0&PlusMinus=N&Rank=N&Season=2015-16&SeasonSegment=&SeasonType=Playoffs&TeamID=" + str(team[0]) + \
                 "&VsConference=&VsDivision="

    response = requests.get(playoffUrl, headers=heads).json()["resultSets"][1]["rowSet"]

    for player in response:
        teams[team][player[1]] = player[7]

    regularUrl = "http://stats.nba.com/stats/teamplayerdashboard?DateFrom=&DateTo=&GameSegment=&LastNGames=0&LeagueID" \
                 "=00&Location=&MeasureType=Base&Month=0&OpponentTeamID=0&Outcome=&PaceAdjust=N&PerMode=PerGame&Perio" \
                 "d=0&PlusMinus=N&Rank=N&Season=2015-16&SeasonSegment=&SeasonType=Regular+Season&TeamID=" + str(team[0]) + \
                 "&VsConference=&VsDivision="

    response = requests.get(regularUrl, headers=heads).json()["resultSets"][1]["rowSet"]
    for player in response:
        if player[3] > 20 and player[1] in teams[team].keys():
            teams[team][player[1]] -= player[7]

    teamDiff = 0
    for min in teams[team].values():
        teamDiff += abs(min)

    ws["A" + str(rowCounter)] = team[1]
    ws["B" + str(rowCounter)] = teamDiff
    rowCounter += 1

wb.save('C:\\Users\\tuvia\\Documents\\Analytics\\Minute Difference.xlsx')
